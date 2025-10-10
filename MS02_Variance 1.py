from openpyxl import load_workbook, Workbook

input_file = "Variance.xlsx"
output_file = "VarianceCommentaryOutputFile.xlsx"

required_sheets = ["SoFP", "SoPL"]

wb_in = load_workbook(input_file, data_only=True)
wb_out = Workbook()

# Remove the default sheet created by openpyxl
if "Sheet" in wb_out.sheetnames:
    wb_out.remove(wb_out["Sheet"])

for sheet_name in required_sheets:
    if sheet_name not in wb_in.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found in the input file.")
        continue

    ws_in = wb_in[sheet_name]
    new_sheet_name = f"{sheet_name}_Variance"
    ws_out = wb_out.create_sheet(title=new_sheet_name)

    # Write headers
    headers = ["FSLI", "CY", "PY", "difference", "difference_percent"]
    ws_out.append(headers)

    # Extract and write data (columns B–F)
    for row in ws_in.iter_rows(min_row=9, min_col=2, max_col=6, values_only=True):
        if all(v is None for v in row):
            continue
            row = list(row)

        row = list(row) 
        if isinstance(row[4], (int, float)) and 0 <= row[4] <= 1:
            row[4] = f"{round(row[4] * 100)}%"

        ws_out.append(row)

wb_out.save(output_file)
print(f"Data extracted and saved to {output_file}")